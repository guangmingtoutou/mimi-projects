# -*- coding: utf-8 -*-
"""批量试卷分析：解析 xlsx 答题数据（含"正确答案"行）、按学生生成报告、打包 zip"""
import re
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from .analyzer import Question, analyze, build_study_plan, grade_question
from .config import REPORT_DIR
from .llm import analyze_paper_llm, llm_available, template_advice
from .report_builder import build_html
from .exporters import html_to_pdf

NAME_KEYS = ["学生姓名", "姓名", "学生", "名字"]
TEACHER_KEYS = ["规划师", "老师", "教师"]
CLASS_KEYS = ["班型", "班级"]
ANSWER_KEY_NAME = "正确答案"

# 表头中的题型标识
TYPE_MARKERS = [
    ("single", ["单项选择题", "单选题"]),
    ("multi", ["多项选择题", "多选题", "不定项"]),
    ("calculation", ["主观题", "计算题", "解答题", "非选择题"]),
    ("experiment", ["实验题"]),
    ("fill", ["填空题"]),
]

QID_RE = re.compile(r"(\d{1,3})\s*[（(]?\s*(\d{1,2})\s*[)）]?")          # 15（2）→ 15,2
SCORE_RE = re.compile(r"(\d{1,3}(?:\.\d)?)\s*分")                        # 5分 / 12分
PARTIAL_RE = re.compile(r"选对但不全[得]\s*(\d{1,2}(?:\.\d)?)\s*分")     # 选对但不全得2分


class XlsxParseError(Exception):
    pass


def _cell_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # 去除 Excel 里不可见字符/换行
    return re.sub(r"[\n\r\t\x00-\x1f]", "", s)


def parse_xlsx(path: Path) -> dict:
    """解析答题表（read_only，限制行数）。
    返回: {students, answer_key, question_ids, question_meta, teachers, teacher_col, class_col}
    question_meta: [{qid, col, qtype, default_score, partial_score}]
    """
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i >= 5000:
            break

    # 找到表头行（含"姓名"）
    header_idx = None
    for i, r in enumerate(rows[:10]):
        if r and any(c is not None and any(k in _cell_str(c) for k in NAME_KEYS) for c in r):
            header_idx = i
            break
    if header_idx is None:
        raise XlsxParseError("未找到含「姓名」的表头行")
    header = rows[header_idx]

    name_col = teacher_col = class_col = None
    for i, c in enumerate(header):
        s = _cell_str(c)
        if not s:
            continue
        if name_col is None and any(k in s for k in NAME_KEYS):
            name_col = i
        elif teacher_col is None and any(k in s for k in TEACHER_KEYS):
            teacher_col = i
        elif class_col is None and any(k in s for k in CLASS_KEYS):
            class_col = i
    if name_col is None:
        raise XlsxParseError("未找到姓名列（请包含「姓名」表头）")

    # 题目列：从表头提取题号与题型/分值信息（分值/题型按“大题”传播给该题区所有小题）
    question_meta = []
    cur_type = "single"
    cur_score = None
    cur_partial = None
    for i, c in enumerate(header):
        s = _cell_str(c)
        if i in (name_col, teacher_col, class_col) or not s:
            continue
        # 大题区段头：更新当前题区类型与分值（如 “一、单项选择题（10题 × 4分 = 40分）…1.（必填）” ）
        for key, markers in TYPE_MARKERS:
            if any(m in s for m in markers):
                cur_type = key
                break
        m = SCORE_RE.search(s)
        if m:
            cur_score = float(m.group(1))
        if cur_type == "multi":
            pm = PARTIAL_RE.search(s)
            if pm:
                cur_partial = float(pm.group(1))
            elif cur_partial is None:
                cur_partial = 2.0
        qid = _extract_qid(s)
        if qid is None:
            continue
        # 主观题没有“每题X分”信息时不给默认分（由试卷/人工配置）
        default_score = cur_score if cur_type != "calculation" else None
        question_meta.append({
            "qid": qid, "col": i, "qtype": cur_type,
            "default_score": default_score, "partial_score": cur_partial if cur_type == "multi" else None,
        })
    if not question_meta:
        raise XlsxParseError("未识别到任何题目列（表头需为“1.（必填）”这类格式）")

    # 数据行 + 正确答案行
    answer_row = None
    students = []
    for r in rows[header_idx + 1:]:
        if not r or all(v is None or _cell_str(v) == "" for v in r):
            continue
        name = _cell_str(r[name_col]) if name_col < len(r) else ""
        if not name:
            continue
        if name == ANSWER_KEY_NAME or "正确答案" in name:
            answer_row = r
            continue
        student = {
            "name": name,
            "teacher": _cell_str(r[teacher_col]) if teacher_col is not None and teacher_col < len(r) else "",
            "class_type": _cell_str(r[class_col]) if class_col is not None and class_col < len(r) else "",
            "answers": {},
        }
        for meta in question_meta:
            v = r[meta["col"]] if meta["col"] < len(r) else None
            student["answers"][meta["qid"]] = _cell_str(v)
        students.append(student)

    if answer_row is None:
        raise XlsxParseError(f"未找到名为「{ANSWER_KEY_NAME}」的标准答案行")
    answer_key = {
        meta["qid"]: (_cell_str(answer_row[meta["col"]]) if meta["col"] < len(answer_row) else "")
        for meta in question_meta
    }
    teachers = sorted({s["teacher"] for s in students if s["teacher"]})
    return {
        "students": students,
        "answer_key": answer_key,
        "question_ids": [m["qid"] for m in question_meta],
        "question_meta": question_meta,
        "teachers": teachers,
        "teacher_col": teacher_col,
        "class_col": class_col,
        "header_idx": header_idx,
    }


def _extract_qid(header_cell: str) -> str | None:
    """从表头单元格提取题号。
    '1.（必填）' → '1'；'15（2）（必填）' → '15（2）'；'三、主观题...15(1)（必填）' → '15(1)'
    """
    s = header_cell.strip()
    if not s or "必填" not in s:
        return None
    # 优先取最后一个题号模式
    m = re.search(r"(\d{1,3})\s*[（(]\s*(\d{1,2})\s*[)）]\s*（必填）", s)
    if m:
        return f"{m.group(1)}({m.group(2)})"
    m = re.search(r"(\d{1,3})\s*[.、．]\s*（必填）", s)
    if m:
        return m.group(1)
    m = re.search(r"(\d{1,3})", s)
    if m:
        return m.group(1)
    return None


def _merge_paper_config(paper_config: list, question_meta: list) -> list:
    """用 xlsx 表头信息补全题目配置（题型/默认分值/多选部分分）
    同时把试卷配置中“大题分值”（如 15=12分）均分到对应小题（15(1)/15(2)/15(3)）"""
    meta_by_qid = {m["qid"]: m for m in question_meta}
    merged = []
    for q in paper_config:
        qid = str(q.get("qid", ""))
        m = meta_by_qid.get(qid)
        entry = dict(q)
        if m:
            entry.setdefault("qtype", m["qtype"])
            if not entry.get("full_score") and m.get("default_score"):
                entry["full_score"] = m["default_score"]
            entry.setdefault("partial_score", m.get("partial_score"))
        merged.append(entry)
    # 小题分值均分：parent(如15)=12分 → 15(1)/15(2)/15(3) 各4分
    parent_by_part = {}
    for q in merged:
        qid = str(q.get("qid", ""))
        mm = re.match(r"^(\d{1,3})\((\d{1,2})\)$", qid)
        if mm:
            parent_by_part.setdefault(mm.group(1), []).append(q)
    for parent_qid, parts in parent_by_part.items():
        if len(parts) < 2:
            continue
        parent = next((q for q in merged if str(q.get("qid")) == parent_qid), None)
        if not parent or not parent.get("full_score"):
            continue
        missing = [p for p in parts if not p.get("full_score")]
        if not missing:
            continue
        parent_score = float(parent["full_score"])
        share = round(parent_score / len(parts), 1)
        assigned = 0.0
        for i, p in enumerate(parts):
            if not p.get("full_score"):
                # 最后一个小题补足差额，保证总和等于大题分值
                if i == len(parts) - 1:
                    p["full_score"] = round(parent_score - assigned, 1)
                else:
                    p["full_score"] = share
                assigned += float(p["full_score"])
        # 大题本身不再参与计分（已拆到小题）
        parent["_skip"] = True
    merged = [q for q in merged if not q.get("_skip")]
    return merged


def run_batch(parsed: dict, teacher: str, paper_config: list, meta: dict, mode: str = "hybrid") -> dict:
    """为指定老师名下的所有学生逐个生成报告，返回 zip 路径与统计信息"""
    students = [s for s in parsed["students"] if s["teacher"] == teacher] if teacher else parsed["students"]
    if not students:
        raise XlsxParseError(f"老师「{teacher}」名下没有学生")

    paper_config = _merge_paper_config(paper_config, parsed.get("question_meta", []))
    qconfig = {str(q["qid"]): q for q in paper_config}
    answer_key = parsed["answer_key"]
    class_type = meta.get("class_type", "目标班")
    use_llm = mode in ("llm", "hybrid") and llm_available()
    global_partial = bool(meta.get("multi_partial", False))

    zip_dir = REPORT_DIR / f"batch_{meta.get('report_id', 'tmp')}"
    zip_dir.mkdir(parents=True, exist_ok=True)

    # 第一步：并行计算所有学生（不写文件，先算排名）
    def _compute_one(s):
        questions = []
        for qid, cfg in qconfig.items():
            full = float(cfg.get("full_score") or 0)
            partial_score = cfg.get("partial_score")
            q = Question(
                qid=qid,
                section_key=cfg.get("section_key", "lixue"),
                qtype=cfg.get("qtype", "single"),
                full_score=full,
                knowledge_point=(cfg.get("knowledge_point") or "").strip(),
                student_answer=s["answers"].get(qid, ""),
                correct_answer=answer_key.get(qid, ""),
                partial_score=float(partial_score) if partial_score else None,
            )
            q.got_score = grade_question(q, multi_partial=global_partial or q.partial_score is not None)
            questions.append(q)
        analysis = analyze(questions)
        if use_llm:
            try:
                advice = analyze_paper_llm(analysis, questions, class_type, [])
            except Exception:
                advice = template_advice(analysis)
        else:
            advice = template_advice(analysis)
        plan = build_study_plan(analysis, class_type)
        name_safe = re.sub(r'[\\/:*?"<>|]', "_", s["name"])
        return {
            "name": s["name"], "name_safe": name_safe, "score": analysis["total_got"],
            "full": analysis["total_full"], "rate": analysis["overall_rate"],
            "analysis": analysis, "advice": advice, "plan": plan,
            "class_type": s.get("class_type") or class_type,
        }

    items = []
    with ThreadPoolExecutor(max_workers=4) as pool:
        for r in pool.map(_compute_one, students):
            items.append(r)
    items.sort(key=lambda x: -x["score"])
    for idx, item in enumerate(items):
        item["rank"] = idx + 1
        item["total_students"] = len(items)

    # 第二步：并行写 HTML（带排名）
    def _write_html(item):
        m = dict(meta)
        m["student"] = item["name"]
        m["teacher"] = teacher
        m["class_type"] = item["class_type"]
        m["rank"] = item["rank"]
        m["total_students"] = item["total_students"]
        m["mode"] = "batch"
        html = build_html(item["analysis"], m, item["advice"], item["plan"], mode="batch")
        html_path = zip_dir / f"{item['name_safe']}_试卷分析报告.html"
        html_path.write_text(html, encoding="utf-8")
        return {**item, "html_path": html_path}

    built = []
    with ThreadPoolExecutor(max_workers=4) as pool:
        for r in pool.map(_write_html, items):
            built.append(r)

    # 第三步：并行导出 PDF（每个学生独立 Edge 进程）
    def _pdf_one(item):
        try:
            pdf_path = zip_dir / f"{item['name_safe']}_试卷分析报告.pdf"
            html_to_pdf(item["html_path"], pdf_path)
            return {**item, "pdf": pdf_path.name}
        except Exception as e:
            print(f"[warn] {item['name']} PDF 导出失败: {e}")
            return {**item, "pdf": None}

    results = []
    with ThreadPoolExecutor(max_workers=4) as pool:
        for r in pool.map(_pdf_one, built):
            results.append({
                "name": r["name"], "score": r["score"], "full": r["full"],
                "rate": r["rate"], "html": r["html_path"].name, "pdf": r["pdf"],
                "rank": r["rank"], "total_students": r["total_students"],
            })

    zip_path = REPORT_DIR / f"batch_{meta.get('report_id', 'tmp')}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in sorted(zip_dir.glob("*.pdf")) + sorted(zip_dir.glob("*.html")):
            zf.write(f, f.name)
    results.sort(key=lambda r: -r["score"])
    return {"zip_path": str(zip_path), "results": results, "count": len(results)}
